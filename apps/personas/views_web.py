"""Vistas web para ficha personal y autogestión de funcionarios."""
from django.views.generic import ListView, DetailView, TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q
from django.db import transaction
from django.core.exceptions import PermissionDenied
from django.conf import settings

from .models import Persona
from .forms import FichaPersonalForm
from apps.laboral.models import DatosLaborales
from apps.academico.models import DatosAcademicos
from apps.ubicacion.models import Domicilio
from apps.auditoria.utils import registrar_auditoria


def build_static_map_url(lat, lng):
    if lat in (None, '') or lng in (None, ''):
        return ''
    if settings.GOOGLE_MAPS_API_KEY:
        return (
            f"https://maps.googleapis.com/maps/api/staticmap?center={lat},{lng}"
            f"&zoom=18&size=900x420&scale=2&maptype=satellite"
            f"&markers=color:red%7C{lat},{lng}&key={settings.GOOGLE_MAPS_API_KEY}"
        )
    return ''





def persona_snapshot(persona):
    if not persona:
        return {}
    dl = getattr(persona, 'datos_laborales', None)
    extra = persona.FICHA_EXTRA or {}
    return {
        'ci_numero': persona.CI_NUMERO,
        'nombres': persona.NOMBRES,
        'apellidos': persona.APELLIDOS,
        'telefono': persona.TELEFONO,
        'email': persona.EMAIL,
        'activo': persona.ACTIVO,
        'celular': extra.get('CELULAR', ''),
        'ciudad': extra.get('CIUDAD', ''),
        'barrio': extra.get('BARRIO', ''),
        'domicilio_actual': extra.get('DOMICILIO_ACTUAL', ''),
        'tipo_vinculo': getattr(dl, 'TIPO_VINCULO', ''),
        'institucion_origen': getattr(dl, 'INSTITUCION_ORIGEN', ''),
        'cargo': getattr(dl, 'CARGO', ''),
        'dependencia': getattr(dl, 'DEPENDENCIA', ''),
        'salario': str(getattr(dl, 'SALARIO', '') or ''),
        'fecha_ingreso': getattr(dl, 'FECHA_INGRESO', None).isoformat() if getattr(dl, 'FECHA_INGRESO', None) else '',
    }

def get_persona_for_user(user):
    """Obtiene la ficha vinculada al usuario; si no existe, intenta vincularla por CI o username."""
    if not user.is_authenticated:
        return None
    persona = Persona.objects.filter(USUARIO=user).first()
    if persona:
        return persona

    possible_ci = [v for v in [getattr(user, 'CI', None), user.username] if v]
    for ci in possible_ci:
        persona = Persona.objects.filter(CI_NUMERO=ci).first()
        if persona and persona.USUARIO_id is None:
            persona.USUARIO = user
            persona.save(update_fields=['USUARIO'])
            return persona
    return None


class RRHHRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.puede_acceder_panel_rrhh

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.info(self.request, 'Accedé desde tu ficha personal.')
            return redirect('mi-ficha')
        return super().handle_no_permission()


class PersonaFormMixin:
    template_name = 'personas/ficha_form.html'

    def _form_context(self, **extra_context):
        ctx = {
            'google_maps_api_key': settings.GOOGLE_MAPS_API_KEY,
            'google_maps_enabled': bool(settings.GOOGLE_MAPS_API_KEY),
        }
        ctx.update(extra_context)
        return ctx

    @transaction.atomic
    def _save_form(self, form, instance=None, usuario=None):
        cd = form.cleaned_data
        persona = instance or Persona()
        persona.USUARIO = usuario or instance.USUARIO if instance and instance.USUARIO else usuario
        persona.CI_NUMERO = cd['CI_NUMERO']
        persona.NOMBRES = cd['NOMBRES']
        persona.APELLIDOS = cd['APELLIDOS']
        persona.FECHA_NACIMIENTO = cd['FECHA_NACIMIENTO']
        persona.ESTADO_CIVIL = cd.get('ESTADO_CIVIL') or 'SOLTERO'
        persona.TELEFONO = cd.get('TELEFONO', '')
        persona.EMAIL = cd.get('EMAIL', '')
        persona.ACTIVO = cd.get('ACTIVO', False)
        if cd.get('FOTO_CARNET'):
            persona.FOTO_CARNET = cd['FOTO_CARNET']
        extra = {
            'CELULAR': cd.get('CELULAR', ''),
            'DOMICILIO_ACTUAL': cd.get('DOMICILIO_ACTUAL', ''),
            'BARRIO': cd.get('BARRIO', ''),
            'CIUDAD': cd.get('CIUDAD', ''),
            'LATITUD': str(cd.get('LATITUD') or ''),
            'LONGITUD': str(cd.get('LONGITUD') or ''),
            'ALERGIAS_ENFERMEDADES': cd.get('ALERGIAS_ENFERMEDADES', ''),
            'GRUPO_SANGUINEO': cd.get('GRUPO_SANGUINEO', ''),
            'TIENE_SEGURO_MEDICO': cd.get('TIENE_SEGURO_MEDICO', False),
            'SANATORIO': cd.get('SANATORIO', ''),
            'TELEFONO_SANATORIO': cd.get('TELEFONO_SANATORIO', ''),
            'EMERGENCIA_NOMBRE_1': cd.get('EMERGENCIA_NOMBRE_1', ''),
            'EMERGENCIA_TELEFONO_1': cd.get('EMERGENCIA_TELEFONO_1', ''),
            'EMERGENCIA_NOMBRE_2': cd.get('EMERGENCIA_NOMBRE_2', ''),
            'EMERGENCIA_TELEFONO_2': cd.get('EMERGENCIA_TELEFONO_2', ''),
            'NOMBRE_PADRE': cd.get('NOMBRE_PADRE', ''),
            'NOMBRE_MADRE': cd.get('NOMBRE_MADRE', ''),
            'NOMBRE_CONYUGE': cd.get('NOMBRE_CONYUGE', ''),
            'OBSERVACIONES_FICHA': cd.get('OBSERVACIONES_FICHA', ''),
            'CROQUIS_REFERENCIA': cd.get('CROQUIS_REFERENCIA', ''),
            'CROQUIS_MAP_URL': build_static_map_url(cd.get('LATITUD'), cd.get('LONGITUD')),
            'FIRMA_ACLARACION': cd.get('FIRMA_ACLARACION', ''),
            'FECHA_FIRMA': cd.get('FECHA_FIRMA').isoformat() if cd.get('FECHA_FIRMA') else '',
            'NUMERO_DGRP': cd.get('NUMERO_DGRP', ''),
            'FECHA_DGRP': cd.get('FECHA_DGRP').isoformat() if cd.get('FECHA_DGRP') else '',
            'CATEGORIA': cd.get('CATEGORIA', ''),
            'INSTITUCION_DESTINO_COMISION': cd.get('INSTITUCION_DESTINO_COMISION', ''),
            'NUMERO_RESOLUCION_COMISION': cd.get('NUMERO_RESOLUCION_COMISION', ''),
            'FECHA_INICIO_COMISION': cd.get('FECHA_INICIO_COMISION').isoformat() if cd.get('FECHA_INICIO_COMISION') else '',
            'FECHA_FIN_COMISION': cd.get('FECHA_FIN_COMISION').isoformat() if cd.get('FECHA_FIN_COMISION') else '',
            'OBSERVACION_COMISION': cd.get('OBSERVACION_COMISION', ''),
            'NIVEL_PRIMARIO': cd.get('NIVEL_PRIMARIO', False),
            'NIVEL_SECUNDARIO': cd.get('NIVEL_SECUNDARIO', False),
            'NIVEL_UNIVERSITARIO': cd.get('NIVEL_UNIVERSITARIO', False),
            'BACHILLER': cd.get('BACHILLER', ''),
            'ANIO_BACHILLER': cd.get('ANIO_BACHILLER'),
            'POSTGRADOS': cd.get('POSTGRADOS', ''),
            'MAESTRIAS': cd.get('MAESTRIAS', ''),
            'DOCTORADOS': cd.get('DOCTORADOS', ''),
            'OTROS_ESTUDIOS': cd.get('OTROS_ESTUDIOS', ''),
        }
        for i in range(1, 7):
            extra[f'HIJO_{i}_NOMBRE'] = cd.get(f'HIJO_{i}_NOMBRE', '')
            extra[f'HIJO_{i}_FECHA'] = cd.get(f'HIJO_{i}_FECHA').isoformat() if cd.get(f'HIJO_{i}_FECHA') else ''
        persona.FICHA_EXTRA = extra
        persona.save()

        if usuario and getattr(usuario, 'CI', None) != persona.CI_NUMERO:
            usuario.CI = persona.CI_NUMERO
            if not usuario.first_name:
                usuario.first_name = persona.NOMBRES
            if not usuario.last_name:
                usuario.last_name = persona.APELLIDOS
            if persona.EMAIL and not usuario.email:
                usuario.email = persona.EMAIL
            usuario.save()

        if cd.get('DOMICILIO_ACTUAL') or cd.get('CIUDAD') or cd.get('BARRIO') or cd.get('LATITUD') or cd.get('LONGITUD'):
            domicilio, _ = Domicilio.objects.get_or_create(persona=persona, ES_ACTUAL=True, defaults={'CIUDAD': cd.get('CIUDAD') or 'No especificada'})
            domicilio.DIRECCION = cd.get('DOMICILIO_ACTUAL', '')
            domicilio.BARRIO = cd.get('BARRIO', '')
            domicilio.CIUDAD = cd.get('CIUDAD', '') or 'No especificada'
            domicilio.LATITUD = cd.get('LATITUD') or None
            domicilio.LONGITUD = cd.get('LONGITUD') or None
            domicilio.save()

        if any(cd.get(k) for k in ['TIPO_VINCULO', 'INSTITUCION_ORIGEN', 'CARGO', 'LUGAR_TRABAJO', 'SALARIO', 'FECHA_INGRESO', 'NUMERO_DECRETO', 'NUMERO_RESOLUCION', 'NUMERO_DGRP', 'FECHA_DGRP', 'INSTITUCION_DESTINO_COMISION', 'NUMERO_RESOLUCION_COMISION', 'FECHA_INICIO_COMISION', 'FECHA_FIN_COMISION', 'OBSERVACION_COMISION']):
            dl, _ = DatosLaborales.objects.get_or_create(persona=persona)
            dl.TIPO_VINCULO = cd.get('TIPO_VINCULO') or 'OTRO'
            dl.INSTITUCION_ORIGEN = cd.get('INSTITUCION_ORIGEN') or 'RUN'
            dl.NUMERO_DECRETO = cd.get('NUMERO_DECRETO', '')
            dl.FECHA_DECRETO = cd.get('FECHA_DECRETO')
            dl.NUMERO_RESOLUCION = cd.get('NUMERO_RESOLUCION', '')
            dl.FECHA_RESOLUCION = cd.get('FECHA_RESOLUCION')
            dl.INSTITUCION_DESTINO_COMISION = cd.get('INSTITUCION_DESTINO_COMISION', '')
            dl.NUMERO_RESOLUCION_COMISION = cd.get('NUMERO_RESOLUCION_COMISION', '')
            dl.FECHA_INICIO_COMISION = cd.get('FECHA_INICIO_COMISION')
            dl.FECHA_FIN_COMISION = cd.get('FECHA_FIN_COMISION')
            dl.OBSERVACION_COMISION = cd.get('OBSERVACION_COMISION', '')
            dl.CARGO = cd.get('CARGO', '')
            dl.DEPENDENCIA = cd.get('LUGAR_TRABAJO', '')
            dl.SALARIO = cd.get('SALARIO')
            dl.FECHA_INGRESO = cd.get('FECHA_INGRESO')
            dl.save()

        if any(cd.get(k) for k in ['CARRERA', 'UNIVERSIDAD', 'ANIO_CARRERA', 'POSTGRADOS', 'MAESTRIAS', 'DOCTORADOS']):
            da, _ = DatosAcademicos.objects.get_or_create(persona=persona)
            if cd.get('DOCTORADOS'):
                da.NIVEL_ACADEMICO = 'DOCTORADO'
            elif cd.get('MAESTRIAS'):
                da.NIVEL_ACADEMICO = 'MAESTRIA'
            elif cd.get('POSTGRADOS'):
                da.NIVEL_ACADEMICO = 'POSTGRADO'
            elif cd.get('NIVEL_UNIVERSITARIO'):
                da.NIVEL_ACADEMICO = 'UNIVERSITARIO'
            elif cd.get('NIVEL_SECUNDARIO'):
                da.NIVEL_ACADEMICO = 'SECUNDARIO'
            elif cd.get('NIVEL_PRIMARIO'):
                da.NIVEL_ACADEMICO = 'PRIMARIO'
            else:
                da.NIVEL_ACADEMICO = 'NINGUNO'
            da.PROFESION = cd.get('CARRERA', '')
            da.UNIVERSIDAD = cd.get('UNIVERSIDAD', '')
            da.ANIO_GRADUACION = cd.get('ANIO_CARRERA')
            da.TIENE_POSTGRADO = bool(cd.get('POSTGRADOS'))
            da.TIENE_MAESTRIA = bool(cd.get('MAESTRIAS'))
            da.TIENE_DOCTORADO = bool(cd.get('DOCTORADOS'))
            da.save()

        return persona


class PersonaListView(RRHHRequiredMixin, ListView):
    model = Persona
    template_name = 'personas/list.html'
    context_object_name = 'object_list'
    paginate_by = 20

    def get_queryset(self):
        queryset = Persona.objects.select_related('USUARIO').select_related('datos_laborales').all()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(CI_NUMERO__icontains=search) |
                Q(NOMBRES__icontains=search) |
                Q(APELLIDOS__icontains=search) |
                Q(TELEFONO__icontains=search) |
                Q(EMAIL__icontains=search)
            )
        estado_civil = self.request.GET.get('ESTADO_CIVIL')
        if estado_civil:
            queryset = queryset.filter(ESTADO_CIVIL=estado_civil)

        activo = self.request.GET.get('ACTIVO')
        if activo == 'true':
            queryset = queryset.filter(ACTIVO=True)
        elif activo == 'false':
            queryset = queryset.filter(ACTIVO=False)

        tipo_vinculo = self.request.GET.get('TIPO_VINCULO')
        if tipo_vinculo:
            queryset = queryset.filter(datos_laborales__TIPO_VINCULO=tipo_vinculo)

        institucion_origen = self.request.GET.get('INSTITUCION_ORIGEN')
        if institucion_origen:
            queryset = queryset.filter(datos_laborales__INSTITUCION_ORIGEN=institucion_origen)

        return queryset.order_by('APELLIDOS', 'NOMBRES')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estado_civil_choices'] = Persona.ESTADO_CIVIL_CHOICES
        context['tipo_vinculo_choices'] = DatosLaborales.TIPO_VINCULO_CHOICES
        context['institucion_origen_choices'] = DatosLaborales.INSTITUCION_ORIGEN_CHOICES
        return context

    def render_to_response(self, context, **response_kwargs):
        export_type = self.request.GET.get('export')
        if export_type == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="funcionarios.csv"'
            response.write('﻿')
            writer = csv.writer(response, delimiter=';')
            writer.writerow(['CI', 'Nombre completo', 'F. Nacimiento', 'Edad', 'Estado civil', 'Teléfono', 'Estado', 'Vínculo', 'Institución de origen', 'Categoría', 'Antigüedad institución de origen', 'Antigüedad RUN'])
            for persona in context['object_list']:
                dl = getattr(persona, 'datos_laborales', None)
                extra = persona.FICHA_EXTRA or {}
                writer.writerow([
                    persona.CI_NUMERO,
                    persona.nombre_completo,
                    persona.FECHA_NACIMIENTO.strftime('%d/%m/%Y') if persona.FECHA_NACIMIENTO else '',
                    persona.edad or '',
                    persona.get_ESTADO_CIVIL_display(),
                    persona.TELEFONO or '-',
                    'Activo' if persona.ACTIVO else 'Inactivo',
                    dl.get_TIPO_VINCULO_display() if dl else '-',
                    dl.institucion_origen_label if dl else '-',
                    extra.get('CATEGORIA', '') or '-',
                    persona.antiguedad_origen or '-',
                    persona.antiguedad_run or '-',
                ])
            return response

        if export_type == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Funcionarios'

            headers = ['CI', 'Nombre completo', 'F. Nacimiento', 'Edad', 'Estado civil', 'Teléfono', 'Estado', 'Vínculo', 'Institución de origen', 'Categoría', 'Antigüedad institución de origen', 'Antigüedad RUN']
            ws.append(headers)

            header_fill = PatternFill(fill_type='solid', fgColor='1F6FEB')
            header_font = Font(color='FFFFFF', bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for persona in context['object_list']:
                dl = getattr(persona, 'datos_laborales', None)
                extra = persona.FICHA_EXTRA or {}
                ws.append([
                    persona.CI_NUMERO,
                    persona.nombre_completo,
                    persona.FECHA_NACIMIENTO.strftime('%d/%m/%Y') if persona.FECHA_NACIMIENTO else '',
                    persona.edad or '',
                    persona.get_ESTADO_CIVIL_display(),
                    persona.TELEFONO or '-',
                    'Activo' if persona.ACTIVO else 'Inactivo',
                    dl.get_TIPO_VINCULO_display() if dl else '-',
                    dl.institucion_origen_label if dl else '-',
                    extra.get('CATEGORIA', '') or '-',
                    persona.antiguedad_origen or '-',
                    persona.antiguedad_run or '-',
                ])

            for column_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        value = str(cell.value or '')
                    except Exception:
                        value = ''
                    max_length = max(max_length, len(value))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="funcionarios.xlsx"'
            return response

        return super().render_to_response(context, **response_kwargs)


class PersonaContextMixin:
    def build_context(self, persona):
        context = {'persona': persona}
        context['extra'] = persona.FICHA_EXTRA or {}
        context['familiares'] = persona.familiares.all() if hasattr(persona, 'familiares') else []
        context['observaciones'] = persona.observaciones.all()[:10] if hasattr(persona, 'observaciones') else []
        context['datos_laborales'] = getattr(persona, 'datos_laborales', None)
        context['datos_academicos'] = getattr(persona, 'datos_academicos', None)
        context['interinatos'] = persona.interinatos.select_related('creado_por').all() if hasattr(persona, 'interinatos') else []
        return context




class PersonaDetailView(RRHHRequiredMixin, PersonaContextMixin, DetailView):
    model = Persona
    template_name = 'personas/detail.html'
    context_object_name = 'persona'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.build_context(self.object))
        context['owner_mode'] = False
        return context


class PersonaPrintView(LoginRequiredMixin, PersonaContextMixin, DetailView):
    model = Persona
    template_name = 'personas/print.html'
    context_object_name = 'persona'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.user.puede_acceder_panel_rrhh:
            return super().dispatch(request, *args, **kwargs)
        own = get_persona_for_user(request.user)
        if not own or own.pk != self.object.pk:
            raise PermissionDenied('No podés imprimir fichas de otros usuarios.')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.build_context(self.object))
        return context


class PersonaCreateView(RRHHRequiredMixin, PersonaFormMixin, View):
    success_url = reverse_lazy('persona-list')

    def get(self, request):
        form = FichaPersonalForm()
        return render(request, self.template_name, self._form_context(form=form, titulo='Nueva Ficha Personal', owner_mode=False))

    def post(self, request):
        form = FichaPersonalForm(request.POST, request.FILES)
        if form.is_valid():
            persona = self._save_form(form)
            registrar_auditoria(
                usuario=request.user,
                accion='CREAR',
                modelo='Persona',
                objeto_id=persona.pk,
                objeto_repr=persona.nombre_completo,
                descripcion=f'Creó la ficha de {persona.nombre_completo}.',
                despues=persona_snapshot(persona),
            )
            messages.success(request, 'Ficha personal creada exitosamente.')
            return redirect('persona-detail', pk=persona.pk)
        messages.error(request, 'Verificá los datos cargados.')
        return render(request, self.template_name, self._form_context(form=form, titulo='Nueva Ficha Personal', owner_mode=False))


class PersonaUpdateView(RRHHRequiredMixin, PersonaFormMixin, View):
    template_name = 'personas/ficha_form.html'

    def get(self, request, pk):
        persona = get_object_or_404(Persona, pk=pk)
        form = FichaPersonalForm(persona=persona)
        return render(request, self.template_name, self._form_context(form=form, titulo='Editar Ficha Personal', persona=persona, owner_mode=False))

    def post(self, request, pk):
        persona = get_object_or_404(Persona, pk=pk)
        form = FichaPersonalForm(request.POST, request.FILES, persona=persona)
        if form.is_valid():
            datos_antes = persona_snapshot(persona)
            persona = self._save_form(form, instance=persona, usuario=persona.USUARIO)
            registrar_auditoria(
                usuario=request.user,
                accion='EDITAR',
                modelo='Persona',
                objeto_id=persona.pk,
                objeto_repr=persona.nombre_completo,
                descripcion=f'Editó la ficha de {persona.nombre_completo}.',
                antes=datos_antes,
                despues=persona_snapshot(persona),
            )
            messages.success(request, 'Ficha personal actualizada exitosamente.')
            return redirect('persona-detail', pk=persona.pk)
        messages.error(request, 'Verificá los datos cargados.')
        return render(request, self.template_name, self._form_context(form=form, titulo='Editar Ficha Personal', persona=persona, owner_mode=False))


class PersonaDeleteView(RRHHRequiredMixin, View):
    def post(self, request, pk):
        persona = get_object_or_404(Persona, pk=pk)
        nombre = persona.nombre_completo if hasattr(persona, 'nombre_completo') else f"{persona.NOMBRES} {persona.APELLIDOS}".strip()
        datos_antes = persona_snapshot(persona)
        registrar_auditoria(
            usuario=request.user,
            accion='ELIMINAR',
            modelo='Persona',
            objeto_id=persona.pk,
            objeto_repr=nombre,
            descripcion=f'Eliminó la ficha de {nombre}.',
            antes=datos_antes,
        )
        persona.delete()
        messages.success(request, f'Registro eliminado correctamente: {nombre}.')
        return redirect('persona-list')


class MiFichaView(LoginRequiredMixin, PersonaContextMixin, TemplateView):
    template_name = 'personas/detail.html'

    def get(self, request, *args, **kwargs):
        persona = get_persona_for_user(request.user)
        if not persona:
            messages.info(request, 'Completá tu ficha personal para continuar.')
            return redirect('mi-ficha-editar')
        context = self.build_context(persona)
        context['owner_mode'] = True
        return render(request, self.template_name, context)


class MiFichaUpdateView(LoginRequiredMixin, PersonaFormMixin, View):
    template_name = 'personas/ficha_form.html'

    def get(self, request):
        persona = get_persona_for_user(request.user)
        form = FichaPersonalForm(persona=persona) if persona else FichaPersonalForm(initial={
            'CI_NUMERO': getattr(request.user, 'CI', ''),
            'NOMBRES': request.user.first_name,
            'APELLIDOS': request.user.last_name,
            'EMAIL': request.user.email,
            'ACTIVO': True,
        })
        return render(request, self.template_name, self._form_context(form=form, titulo='Mi Ficha Personal', persona=persona, owner_mode=True))

    def post(self, request):
        persona = get_persona_for_user(request.user)
        form = FichaPersonalForm(request.POST, request.FILES, persona=persona)
        if form.is_valid():
            datos_antes = persona_snapshot(persona) if persona else {}
            persona = self._save_form(form, instance=persona, usuario=request.user)
            accion = 'EDITAR' if datos_antes else 'CREAR'
            descripcion = 'Actualizó su ficha personal.' if datos_antes else 'Completó su ficha personal.'
            registrar_auditoria(
                usuario=request.user,
                accion=accion,
                modelo='Persona',
                objeto_id=persona.pk,
                objeto_repr=persona.nombre_completo,
                descripcion=descripcion,
                antes=datos_antes or None,
                despues=persona_snapshot(persona),
            )
            messages.success(request, 'Tus datos fueron guardados correctamente.')
            return redirect('mi-ficha')
        messages.error(request, 'Verificá los datos cargados.')
        return render(request, self.template_name, self._form_context(form=form, titulo='Mi Ficha Personal', persona=persona, owner_mode=True))


class MiFichaPrintView(LoginRequiredMixin, PersonaContextMixin, TemplateView):
    template_name = 'personas/print.html'

    def get(self, request, *args, **kwargs):
        persona = get_persona_for_user(request.user)
        if not persona:
            messages.info(request, 'Primero debés completar tu ficha personal.')
            return redirect('mi-ficha-editar')
        context = self.build_context(persona)
        return render(request, self.template_name, context)


class MapaPersonasView(RRHHRequiredMixin, TemplateView):
    template_name = 'personas/mapa.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ciudades = Domicilio.objects.values('CIUDAD').distinct().order_by('CIUDAD')
        context['ciudades'] = ciudades
        return context
